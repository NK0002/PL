import cv2
import os
from deepface import DeepFace
from openpyxl import load_workbook, workbook

data_dir = "BANCO DE DADOS"

imagem = cv2.imread("screenshot.jpg")
dfs = DeepFace.find(img_path="screenshot.jpg",
                    db_path="C:/Users/di056649/PycharmProjects/pythonProject/BANCO DE DADOS")
print(dfs)
if len(dfs) > 0:
    wb = load_workbook('C:/Users/di056649/Desktop/Pasta1.xlsx')
    planilha = wb.active
    linha_cabecalho = planilha[1]
    ultima_coluna = len(linha_cabecalho) + 1
    ultima_linha = 2
    for file in os.listdir('C:/Users/di056649/PycharmProjects/pythonProject/BANCO DE DADOS'):
        if file == 'representations_vgg_face.pkl':
            break
        else:
            result = DeepFace.verify(imagem, f'C:/Users/di056649/PycharmProjects/pythonProject/BANCO DE DADOS/{file}')
            if result['verified']:
                print(result)
                print(file)
                novo_valor = file.split(".")
                nome = novo_valor[0]
                planilha.cell(row=ultima_linha, column=ultima_coluna, value=nome)
                wb.save('C:/Users/di056649/Desktop/Pasta1.xlsx')
                ultima_linha = ultima_linha + 1
            else:
                print('kkkk')
                novo_valor = file.split(".")
                nome = novo_valor[0] + " Faltou"
                planilha.cell(row=ultima_linha, column=ultima_coluna, value=nome)
                wb.save('C:/Users/di056649/Desktop/Pasta1.xlsx')
                ultima_linha = ultima_linha + 1







